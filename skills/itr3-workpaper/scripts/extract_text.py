#!/usr/bin/env python3
"""
Dumps readable content of client source documents to stdout as plain text.

PDF/images: use the Read tool natively — do not run this script on them.
Excel (.xlsx/.xls) and Word (.docx): use this script.

Usage:
    python3 extract_text.py <file> [file...]
"""
import sys
import os


def dump_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        print(f"\n===== SHEET: {ws.title} =====")
        for row in ws.iter_rows(values_only=True):
            if all(v is None for v in row):
                continue
            print(" | ".join("" if v is None else str(v) for v in row))


def dump_docx(path):
    from docx import Document
    doc = Document(path)
    for element in doc.element.body:
        tag = element.tag.split("}")[-1]
        if tag == "p":
            texts = element.findall(
                ".//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t"
            )
            text = "".join(t.text or "" for t in texts)
            if text.strip():
                print(text)
        elif tag == "tbl":
            print("--- TABLE ---")
            for tr in element.findall(
                "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tr"
            ):
                cells = tr.findall(
                    "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tc"
                )
                row_texts = []
                for tc in cells:
                    texts = tc.findall(
                        ".//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t"
                    )
                    row_texts.append("".join(t.text or "" for t in texts))
                print(" | ".join(row_texts))
            print("--- END TABLE ---")


def dump_one(path):
    print(f"\n########## FILE: {path} ##########")
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls"):
        dump_xlsx(path)
    elif ext == ".docx":
        dump_docx(path)
    elif ext in (".csv", ".txt", ".json", ".md"):
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            print(f.read())
    else:
        print(
            f"Unsupported extension '{ext}'. For .pdf or image files, use the Read tool directly.",
            file=sys.stderr,
        )
        return 1
    return 0


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 extract_text.py <file> [file...]", file=sys.stderr)
        sys.exit(1)
    rc = 0
    for path in sys.argv[1:]:
        if not os.path.isfile(path):
            print(f"Not a file: {path}", file=sys.stderr)
            rc = 1
            continue
        rc = dump_one(path) or rc
    sys.exit(rc)


if __name__ == "__main__":
    main()
